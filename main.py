import pandas as pd


fundamentals_df = pd.read_excel('master.xlsx',
                                sheet_name="fundamentals")
prices_df = pd.read_excel('master.xlsx',
                          sheet_name="prices")
prices_split_adjusted_df = pd.read_excel('master.xlsx',
                                         sheet_name="prices-split-adjusted")
securities_df = pd.read_excel('master.xlsx',
                              sheet_name="securities")

#print(fundamentals_df.head())
#print(fundamentals_df.tail())

# creating fundamentals condensed dataframe
fundamentals_condensed_df = pd.DataFrame()
id = fundamentals_df.iloc[:,0]
fundamentals_condensed_df['ID'] = id.copy()
ticker_symbol = fundamentals_df.iloc[:,1]
fundamentals_condensed_df['Ticker Symbol'] = ticker_symbol.copy()
period_ending = fundamentals_df.iloc[:,2]
fundamentals_condensed_df['Period Ending'] = period_ending.copy()
accounts_payable = fundamentals_df.iloc[:,3]
fundamentals_condensed_df['Accounts Payable'] = accounts_payable.copy()
accounts_receivable = fundamentals_df.iloc[:,4]
fundamentals_condensed_df['Accounts Receivable'] = accounts_receivable.copy()
gross_profit = fundamentals_df.iloc[:,25]
fundamentals_condensed_df['Gross Profit'] = gross_profit.copy()
intangible_assets = fundamentals_df.iloc[:,27]
fundamentals_condensed_df['Intangible Assets'] = intangible_assets.copy()
interest_expense = fundamentals_df.iloc[:,28]
fundamentals_condensed_df['Interest Expense'] = interest_expense.copy()
investments = fundamentals_df.iloc[:,30]
fundamentals_condensed_df['Investments'] = investments.copy()
liabilities = fundamentals_df.iloc[:,31]
fundamentals_condensed_df['Liabilities'] = liabilities.copy()
long_term_debt = fundamentals_df.iloc[:,32]
fundamentals_condensed_df['Long-Term Debt'] = long_term_debt.copy()
long_term_investments = fundamentals_df.iloc[:,33]
fundamentals_condensed_df['Long-Term Investments'] = long_term_investments.copy()
minority_interest = fundamentals_df.iloc[:,34]
fundamentals_condensed_df['Minority Interest'] = minority_interest.copy()
for_year = fundamentals_df.iloc[:,2] # period ending and will slice the month and day as we only want the year value
fundamentals_condensed_df['For Year'] = for_year.copy()
earnings_per_share = fundamentals_df.iloc[:,77]
fundamentals_condensed_df['Earnings Per Share'] = earnings_per_share.copy()
estimated_shares_outstanding = fundamentals_df.iloc[:,78]
fundamentals_condensed_df['Estimated Shares Outstanding'] = estimated_shares_outstanding.copy()


# remove timestamp from period ending values
fundamentals_condensed_df['Period Ending'] = \
    fundamentals_condensed_df['Period Ending'].astype(str).str[:11] #convert to string format and remove timestamp

#remove month and year from for year values
fundamentals_condensed_df['For Year'] = \
    fundamentals_condensed_df['Period Ending'].astype(str).str[:4] #convert to string and grab the year value only

# created new workbook containing fundamentals condensed dataframe
fundamentals_condensed_df.to_excel('fundamentals_condensed_df.xlsx', index=False)

# creating earnings dataframe from fundamentals condensed dataframe
earnings_df = pd.DataFrame()
id = fundamentals_condensed_df.iloc[:,0]
earnings_df['ID'] = id.copy()
ticker_symbol = fundamentals_condensed_df.iloc[:,1]
earnings_df['Ticker Symbol'] = ticker_symbol.copy()
for_year = fundamentals_condensed_df.iloc[:,13]
earnings_df['For Year'] = for_year.copy()
earnings_per_share = fundamentals_condensed_df.iloc[:,14]
earnings_df['Earnings Per Share'] = earnings_per_share.copy()
estimated_shares_outstanding = fundamentals_condensed_df.iloc[:,15]
earnings_df['Estimated Earnings'] = estimated_shares_outstanding.copy()

#creating new earnings per share plus minus column
earnings_per_share_plus_minus = []

for value in earnings_df['Earnings Per Share']:
    try:
        if value >= 0:
            earnings_per_share_plus_minus.append('(+)')
        elif value < 0:
            earnings_per_share_plus_minus.append('(-)')
        else:
            earnings_per_share_plus_minus.append('N/A')
    except:
        print('Error. Cannot append plus/minus value(s).')

earnings_df['Earnings Per Share +/-'] = earnings_per_share_plus_minus

#creating new estimated earnings grade column
estimated_earnings_status = []

for value in earnings_df['Estimated Earnings']:
    try:
        if value >= 500000000:
            estimated_earnings_status.append('Excellent')
        elif value >= 25000000 and value <= 499999999.99:
            estimated_earnings_status.append('Solid')
        elif value >= 1 and value <= 249999999.99:
            estimated_earnings_status.append('Positive')
        elif value == 0:
            estimated_earnings_status.append('Even')
        elif value < 0:
            estimated_earnings_status.append('Failing')
        else:
            estimated_earnings_status.append('N/A')
    except:
        print('Error. Cannot append estimated earnings status values.')

earnings_df['Estimated Earnings Grade'] = estimated_earnings_status

# created new workbook containing fundamentals condensed dataframe
earnings_df.to_excel('earnings_df.xlsx', index=False)


# PIVOT TABLES

#earnings per share pivot table
earnings_per_share_pivot_table = pd.pivot_table(
    earnings_df,
    index='Ticker Symbol',
    columns='For Year',
    values='Earnings Per Share',
    aggfunc='sum'
)

earnings_per_share_pivot_table.to_excel('earnings_per_share_pivot_table.xlsx')

# CONDITIONAL FORMATTING
from openpyxl import load_workbook, formatting, styles

#earnings per share pivot table, cells highlighted with certain colors
wb = load_workbook("earnings_per_share_pivot_table.xlsx")
ws = wb.active
red_color = 'ffc7ce'
red_fill = styles.PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
for row in range(1,10):
    ws.cell(row=row, column=1, value=row-5)
    ws.cell(row=row, column=2, value=row-5)
ws.conditional_formatting.add('A1:A10',
                              formatting.rule.CellIsRule(operator='lessThan', formula=['0'], fill=red_fill)
                              )
ws.conditional_formatting.add('B1:B10',
                              formatting.rule.CellIsRule(operator='lessThan', formula=['0'], fill=red_fill)
                              )
wb.save("earnings_per_share_pivot_table.xlsx")